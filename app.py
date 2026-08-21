import base64
import csv
import io
import json
import os
import shutil
import sqlite3
import uuid
from datetime import date, datetime
from functools import wraps
from pathlib import Path

try:
    import face_recognition
    HAS_FACE_RECOGNITION = True
except Exception:
    face_recognition = None
    HAS_FACE_RECOGNITION = False

import numpy as np
from flask import Flask, flash, g, jsonify, redirect, render_template, request, session, url_for
from PIL import Image
from werkzeug.security import check_password_hash, generate_password_hash


BASE_DIR = Path(__file__).resolve().parent
IS_VERCEL = bool(os.environ.get("VERCEL"))

if IS_VERCEL:
    DATA_DIR = Path("/tmp/data")
    UPLOAD_DIR = Path("/tmp/uploads")
    DATABASE_PATH = DATA_DIR / "college_erp.db"
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    
    seed_db = BASE_DIR / "data" / "college_erp.db"
    if seed_db.exists() and not DATABASE_PATH.exists():
        try:
            shutil.copy2(seed_db, DATABASE_PATH)
        except Exception:
            pass
else:
    DATA_DIR = BASE_DIR / "data"
    UPLOAD_DIR = BASE_DIR / "static" / "uploads"
    DATABASE_PATH = DATA_DIR / "college_erp.db"
    DATA_DIR.mkdir(exist_ok=True)
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

DEFAULT_ADMIN_USERNAME = os.environ.get("ADMIN_USERNAME", "admin")
DEFAULT_ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "admin123")
MATCH_THRESHOLD = 0.47
RESULT_IMPORT_REQUIRED = {
    "enrollment_no",
    "assessment_name",
    "course_code",
    "course_title",
    "score",
    "max_score",
}
RESULT_IMPORT_HEADER_MAP = {
    "enrollment_no": "enrollment_no",
    "enrollment": "enrollment_no",
    "enrollment_number": "enrollment_no",
    "student_code": "enrollment_no",
    "student_id": "enrollment_no",
    "roll_no": "enrollment_no",
    "roll_number": "enrollment_no",
    "assessment_name": "assessment_name",
    "assessment": "assessment_name",
    "exam": "assessment_name",
    "exam_name": "assessment_name",
    "test_name": "assessment_name",
    "course_code": "course_code",
    "subject_code": "course_code",
    "paper_code": "course_code",
    "course_title": "course_title",
    "course_name": "course_title",
    "subject": "course_title",
    "subject_name": "course_title",
    "score": "score",
    "marks": "score",
    "marks_obtained": "score",
    "obtained_marks": "score",
    "max_score": "max_score",
    "max_marks": "max_score",
    "total_marks": "max_score",
    "out_of": "max_score",
}

DATA_DIR.mkdir(exist_ok=True)
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "dev-secret-key-change-me")
app.config["MAX_CONTENT_LENGTH"] = 8 * 1024 * 1024


SCHEDULE_SEED = [
    ("Computer Science", "B.Tech CSE", "Semester 2", "CS201", "Data Structures", "Dr. Meera Iyer", "Block A-302", "Saturday", "08:30", "09:20"),
    ("Computer Science", "B.Tech CSE", "Semester 4", "CS402", "Operating Systems", "Prof. Arjun Sen", "Block A-401", "Saturday", "09:30", "10:20"),
    ("Electronics", "B.Tech ECE", "Semester 6", "EC611", "VLSI Design", "Dr. Ravi Menon", "Lab E-2", "Saturday", "10:30", "11:20"),
    ("Management", "BBA", "Semester 2", "MG204", "Business Statistics", "Prof. Nisha Kapoor", "Block C-112", "Saturday", "11:30", "12:20"),
    ("Commerce", "B.Com", "Semester 4", "BC412", "Corporate Accounting", "Prof. Kavya Shah", "Block D-210", "Saturday", "13:00", "13:50"),
    ("Computer Applications", "BCA", "Semester 6", "CA603", "Cloud Computing", "Dr. Tanish Verma", "Tech Lab-3", "Saturday", "14:00", "14:50"),
]

NOTICE_SEED = [
    ("Internal assessment moderation", "Department coordinators should complete moderation sheets before marks are published to the student portal.", "Academics"),
    ("Hostel fee follow-up", "Accounts office should contact students with pending hostel balances before semester registration closes.", "Finance"),
    ("Biometric gate calibration", "Front gate terminals should be cleaned before first lecture for reliable face attendance capture.", "Operations"),
    ("Placement readiness workshop", "Final-year students should report to Seminar Hall 2 for the campus placement briefing this week.", "Students"),
]


def normalize_text(value) -> str:
    return " ".join(str(value or "").split())


def normalize_enrollment_no(value) -> str:
    return normalize_text(value).upper()


def normalize_course_code(value) -> str:
    return normalize_text(value).upper()


def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DATABASE_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON")
    return g.db


@app.teardown_appcontext
def close_db(_error=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def seed_reference_data(db):
    schedule_count = db.execute("SELECT COUNT(*) AS count FROM course_schedule").fetchone()["count"]
    if schedule_count == 0:
        db.executemany(
            """
            INSERT INTO course_schedule (
                department, program_name, semester, course_code, course_title,
                faculty_name, room_no, weekday, start_time, end_time
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            SCHEDULE_SEED,
        )

    notice_count = db.execute("SELECT COUNT(*) AS count FROM notices").fetchone()["count"]
    if notice_count == 0:
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        db.executemany(
            "INSERT INTO notices (title, body, audience, created_at) VALUES (?, ?, ?, ?)",
            [(title, body, audience, now) for title, body, audience in NOTICE_SEED],
        )


def seed_admin_user(db):
    admin = db.execute("SELECT id FROM admin_users WHERE username = ?", (DEFAULT_ADMIN_USERNAME,)).fetchone()
    if admin:
        return

    db.execute(
        """
        INSERT INTO admin_users (username, full_name, password_hash, created_at)
        VALUES (?, ?, ?, ?)
        """,
        (
            DEFAULT_ADMIN_USERNAME,
            "College Administrator",
            generate_password_hash(DEFAULT_ADMIN_PASSWORD),
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        ),
    )


def init_db():
    schema = """
    CREATE TABLE IF NOT EXISTS students (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        enrollment_no TEXT NOT NULL UNIQUE,
        full_name TEXT NOT NULL,
        department TEXT NOT NULL,
        program_name TEXT NOT NULL,
        semester TEXT NOT NULL,
        batch_year TEXT NOT NULL,
        emergency_contact_name TEXT NOT NULL,
        emergency_contact_phone TEXT NOT NULL,
        email TEXT,
        dob TEXT,
        address TEXT,
        status TEXT NOT NULL DEFAULT 'Active',
        image_path TEXT NOT NULL,
        face_encoding TEXT NOT NULL,
        portal_password_hash TEXT NOT NULL DEFAULT '',
        created_at TEXT NOT NULL
    );

    CREATE TABLE IF NOT EXISTS attendance_logs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        student_id INTEGER NOT NULL,
        attendance_date TEXT NOT NULL,
        status TEXT NOT NULL,
        confidence REAL NOT NULL,
        logged_at TEXT NOT NULL,
        UNIQUE(student_id, attendance_date),
        FOREIGN KEY (student_id) REFERENCES students(id)
    );

    CREATE TABLE IF NOT EXISTS course_results (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        student_id INTEGER NOT NULL,
        assessment_name TEXT NOT NULL,
        course_code TEXT NOT NULL,
        course_title TEXT NOT NULL,
        score REAL NOT NULL,
        max_score REAL NOT NULL,
        grade TEXT NOT NULL,
        updated_at TEXT NOT NULL,
        FOREIGN KEY (student_id) REFERENCES students(id)
    );

    CREATE TABLE IF NOT EXISTS fee_records (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        student_id INTEGER NOT NULL,
        fee_head TEXT NOT NULL,
        semester_label TEXT NOT NULL,
        amount_due REAL NOT NULL,
        amount_paid REAL NOT NULL,
        status TEXT NOT NULL,
        due_date TEXT NOT NULL,
        updated_at TEXT NOT NULL,
        FOREIGN KEY (student_id) REFERENCES students(id)
    );

    CREATE TABLE IF NOT EXISTS course_schedule (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        department TEXT NOT NULL,
        program_name TEXT NOT NULL,
        semester TEXT NOT NULL,
        course_code TEXT NOT NULL,
        course_title TEXT NOT NULL,
        faculty_name TEXT NOT NULL,
        room_no TEXT NOT NULL,
        weekday TEXT NOT NULL,
        start_time TEXT NOT NULL,
        end_time TEXT NOT NULL
    );

    CREATE TABLE IF NOT EXISTS notices (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        title TEXT NOT NULL,
        body TEXT NOT NULL,
        audience TEXT NOT NULL,
        created_at TEXT NOT NULL
    );

    CREATE TABLE IF NOT EXISTS admin_users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT NOT NULL UNIQUE,
        full_name TEXT NOT NULL,
        password_hash TEXT NOT NULL,
        created_at TEXT NOT NULL
    );
    """

    with sqlite3.connect(DATABASE_PATH) as db:
        db.row_factory = sqlite3.Row
        db.execute("PRAGMA foreign_keys = ON")
        db.executescript(schema)
        db.execute(
            """
            DELETE FROM course_results
            WHERE id NOT IN (
                SELECT MAX(id)
                FROM course_results
                GROUP BY student_id, assessment_name, course_code
            )
            """
        )
        db.execute(
            "CREATE UNIQUE INDEX IF NOT EXISTS idx_course_results_unique ON course_results (student_id, assessment_name, course_code)"
        )
        db.execute("CREATE INDEX IF NOT EXISTS idx_students_enrollment_lookup ON students (enrollment_no)")
        db.execute(
            "CREATE INDEX IF NOT EXISTS idx_attendance_logs_lookup ON attendance_logs (student_id, attendance_date)"
        )
        seed_reference_data(db)
        seed_admin_user(db)
        db.commit()


def decode_data_url(data_url: str) -> Image.Image:
    if not data_url or "," not in data_url:
        raise ValueError("Invalid image payload. Capture the image again.")

    encoded = data_url.split(",", 1)[1]
    try:
        raw = base64.b64decode(encoded)
    except ValueError as exc:
        raise ValueError("Image payload could not be decoded.") from exc

    return Image.open(io.BytesIO(raw))


def save_face_image(uploaded_file, captured_photo: str, prefix: str = "student") -> str:
    if uploaded_file and uploaded_file.filename:
        image = Image.open(uploaded_file.stream)
    elif captured_photo:
        image = decode_data_url(captured_photo)
    else:
        raise ValueError("Upload a clear face photo or capture one from the camera.")

    image = image.convert("RGB")
    image.thumbnail((1600, 1600))
    filename = f"{prefix}-{datetime.now().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:8]}.jpg"
    target_path = UPLOAD_DIR / filename
    image.save(target_path, format="JPEG", quality=92)
    return f"uploads/{filename}"


def remove_saved_image(relative_path: str):
    if not relative_path:
        return
    absolute_path = BASE_DIR / "static" / relative_path
    if absolute_path.exists():
        absolute_path.unlink()


def extract_face_encoding(image_path: str | None = None, image_data_url: str | None = None):
    if image_path:
        img_obj = Image.open(BASE_DIR / "static" / image_path).convert("RGB")
    elif image_data_url:
        img_obj = decode_data_url(image_data_url).convert("RGB")
    else:
        raise ValueError("No image was supplied for facial recognition.")

    image_array = np.array(img_obj)

    if HAS_FACE_RECOGNITION and face_recognition is not None:
        locations = face_recognition.face_locations(image_array, model="hog")
        if not locations:
            raise ValueError("No face was detected. Use a clear front-facing photo.")
        if len(locations) > 1:
            raise ValueError("Multiple faces detected. Use an image with only one student.")

        encodings = face_recognition.face_encodings(image_array, known_face_locations=locations)
        if not encodings:
            raise ValueError("Face encoding failed. Try a sharper image.")
        return encodings[0]
    else:
        # Fallback normalized 128D feature vector for environments without dlib
        small_img = img_obj.resize((16, 8))
        flat = np.array(small_img, dtype=np.float64).flatten()
        norm = np.linalg.norm(flat)
        if norm > 0:
            flat = flat / norm
        return flat[:128]



def serialize_face_encoding(encoding) -> str:
    return json.dumps(np.asarray(encoding, dtype=np.float64).tolist())


def parse_face_encoding(encoded_text: str):
    return np.asarray(json.loads(encoded_text), dtype=np.float64)


def normalize_import_header(value) -> str:
    parts = []
    previous_separator = False
    for character in str(value or "").strip().lower():
        if character.isalnum():
            parts.append(character)
            previous_separator = False
        elif not previous_separator:
            parts.append("_")
            previous_separator = True
    return "".join(parts).strip("_")


def canonical_result_header(value) -> str:
    return RESULT_IMPORT_HEADER_MAP.get(normalize_import_header(value), "")


def iter_result_upload_rows(uploaded_file):
    filename = (uploaded_file.filename or "").lower()
    if filename.endswith(".csv"):
        uploaded_file.stream.seek(0)
        content = uploaded_file.stream.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(content))
        if not reader.fieldnames:
            raise ValueError("The uploaded CSV file is missing a header row.")
        headers = [canonical_result_header(header) for header in reader.fieldnames]
        missing = RESULT_IMPORT_REQUIRED - set(headers)
        if missing:
            raise ValueError(f"Missing required columns: {', '.join(sorted(missing))}.")
        for row_number, row in enumerate(reader, start=2):
            normalized = {canonical_result_header(key): row.get(key) for key in reader.fieldnames}
            yield row_number, normalized
        return

    if filename.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError("Excel import requires openpyxl. Install requirements and try again.") from exc

        uploaded_file.stream.seek(0)
        workbook = load_workbook(uploaded_file.stream, read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            workbook.close()
            raise ValueError("The uploaded Excel file is empty.")
        canonical_headers = [canonical_result_header(header) for header in headers]
        missing = RESULT_IMPORT_REQUIRED - set(canonical_headers)
        if missing:
            workbook.close()
            raise ValueError(f"Missing required columns: {', '.join(sorted(missing))}.")
        try:
            for row_number, values in enumerate(rows, start=2):
                normalized = {}
                for index, header in enumerate(canonical_headers):
                    if header:
                        normalized[header] = values[index] if index < len(values) else None
                yield row_number, normalized
        finally:
            workbook.close()
        return

    raise ValueError("Upload a .xlsx or .csv marks file.")


def parse_result_upload(uploaded_file):
    parsed_rows = []
    skipped_blank_rows = 0

    for row_number, row in iter_result_upload_rows(uploaded_file):
        if not any(normalize_text(value) for value in row.values()):
            skipped_blank_rows += 1
            continue

        enrollment_no = normalize_enrollment_no(row.get("enrollment_no"))
        assessment_name = normalize_text(row.get("assessment_name"))
        course_code = normalize_course_code(row.get("course_code"))
        course_title = normalize_text(row.get("course_title"))

        if not enrollment_no:
            raise ValueError(f"Row {row_number}: enrollment number is required.")
        if not assessment_name:
            raise ValueError(f"Row {row_number}: assessment name is required.")
        if not course_code:
            raise ValueError(f"Row {row_number}: course code is required.")
        if not course_title:
            raise ValueError(f"Row {row_number}: course title is required.")

        try:
            score = float(row.get("score"))
            max_score = float(row.get("max_score"))
        except (TypeError, ValueError) as exc:
            raise ValueError(f"Row {row_number}: score and max score must be numeric.") from exc

        if score < 0 or max_score <= 0 or score > max_score:
            raise ValueError(f"Row {row_number}: score must be between 0 and max score.")

        parsed_rows.append(
            {
                "enrollment_no": enrollment_no,
                "assessment_name": assessment_name,
                "course_code": course_code,
                "course_title": course_title,
                "score": score,
                "max_score": max_score,
            }
        )

    return parsed_rows, skipped_blank_rows


def grade_from_percent(percent: float) -> str:
    if percent >= 90:
        return "A+"
    if percent >= 80:
        return "A"
    if percent >= 70:
        return "B+"
    if percent >= 60:
        return "B"
    if percent >= 50:
        return "C"
    return "D"


def fee_status(amount_due: float, amount_paid: float, due_date: str) -> str:
    if amount_paid >= amount_due:
        return "Paid"
    if due_date and due_date < date.today().isoformat():
        return "Overdue"
    if amount_paid > 0:
        return "Part Paid"
    return "Pending"


def upsert_course_results(db, result_rows):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    payload = []
    for row in result_rows:
        score = float(row["score"])
        max_score = float(row["max_score"])
        percent = (score * 100.0) / max_score
        payload.append(
            (
                int(row["student_id"]),
                normalize_text(row["assessment_name"]),
                normalize_course_code(row["course_code"]),
                normalize_text(row["course_title"]),
                score,
                max_score,
                grade_from_percent(percent),
                timestamp,
            )
        )

    db.executemany(
        """
        INSERT INTO course_results (
            student_id, assessment_name, course_code, course_title,
            score, max_score, grade, updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(student_id, assessment_name, course_code)
        DO UPDATE SET
            course_title = excluded.course_title,
            score = excluded.score,
            max_score = excluded.max_score,
            grade = excluded.grade,
            updated_at = excluded.updated_at
        """,
        payload,
    )
    return len(payload)


def redirect_for_role():
    role = session.get("role")
    if role == "admin":
        return redirect(url_for("dashboard"))
    if role == "student":
        return redirect(url_for("student_panel"))
    return redirect(url_for("login"))


def admin_required(view_func):
    @wraps(view_func)
    def wrapped_view(*args, **kwargs):
        if session.get("role") != "admin":
            flash("Please sign in as admin to continue.", "error")
            return redirect(url_for("login"))
        return view_func(*args, **kwargs)

    return wrapped_view


def student_required(view_func):
    @wraps(view_func)
    def wrapped_view(*args, **kwargs):
        if session.get("role") != "student":
            flash("Please sign in as student to continue.", "error")
            return redirect(url_for("login"))
        return view_func(*args, **kwargs)

    return wrapped_view


def verify_student_password(student, password: str) -> bool:
    password_hash = (student["portal_password_hash"] or "").strip()
    if password_hash:
        return check_password_hash(password_hash, password)
    return (student["emergency_contact_phone"] or "") == password


@app.context_processor
def inject_user_context():
    return {
        "current_user_role": session.get("role"),
        "current_user_name": session.get("display_name"),
    }


def get_dashboard_metrics():
    db = get_db()
    total_students = db.execute("SELECT COUNT(*) AS count FROM students").fetchone()["count"]
    active_departments = db.execute("SELECT COUNT(DISTINCT department) AS count FROM students").fetchone()["count"]
    active_programs = db.execute("SELECT COUNT(DISTINCT program_name) AS count FROM students").fetchone()["count"]
    present_today = db.execute(
        "SELECT COUNT(*) AS count FROM attendance_logs WHERE attendance_date = ?",
        (date.today().isoformat(),),
    ).fetchone()["count"]
    avg_score = db.execute(
        "SELECT ROUND(AVG((score * 100.0) / max_score), 1) AS value FROM course_results"
    ).fetchone()["value"]
    finance = db.execute(
        "SELECT COALESCE(SUM(amount_due), 0) AS total_due, COALESCE(SUM(amount_paid), 0) AS total_paid FROM fee_records"
    ).fetchone()

    total_due = round(finance["total_due"], 2)
    total_paid = round(finance["total_paid"], 2)
    attendance_rate = round((present_today * 100.0 / total_students), 1) if total_students else 0

    return {
        "total_students": total_students,
        "active_departments": active_departments,
        "active_programs": active_programs,
        "present_today": present_today,
        "attendance_rate": attendance_rate,
        "average_score": round(avg_score or 0, 1),
        "total_due": total_due,
        "total_paid": total_paid,
        "outstanding": round(total_due - total_paid, 2),
    }
    total_paid = round(finance["total_paid"], 2)
    attendance_rate = round((present_today * 100.0 / total_students), 1) if total_students else 0

    return {
        "total_students": total_students,
        "active_departments": active_departments,
        "active_programs": active_programs,
        "present_today": present_today,
        "attendance_rate": attendance_rate,
        "average_score": round(avg_score or 0, 1),
        "total_due": total_due,
        "total_paid": total_paid,
        "outstanding": round(total_due - total_paid, 2),
    }


def fetch_recent_attendance(limit: int, student_id: int | None = None):
    db = get_db()
    query = """
        SELECT a.attendance_date,
               a.status,
               ROUND(a.confidence, 1) AS confidence,
               a.logged_at,
               s.full_name,
               s.enrollment_no,
               s.department,
               s.program_name,
               s.semester
        FROM attendance_logs a
        JOIN students s ON s.id = a.student_id
    """
    params = []
    if student_id is not None:
        query += " WHERE a.student_id = ?"
        params.append(student_id)
    query += " ORDER BY a.logged_at DESC LIMIT ?"
    params.append(limit)
    return db.execute(query, params).fetchall()


def fetch_students():
    db = get_db()
    return db.execute(
        """
        SELECT s.*,
               COALESCE(SUM(f.amount_due - f.amount_paid), 0) AS balance_due,
               COALESCE(
                   ROUND((SELECT AVG((cr.score * 100.0) / cr.max_score)
                          FROM course_results cr
                          WHERE cr.student_id = s.id), 1),
                   0
               ) AS academic_average
        FROM students s
        LEFT JOIN fee_records f ON f.student_id = s.id
        GROUP BY s.id
        ORDER BY s.created_at DESC
        """
    ).fetchall()


def fetch_student_options():
    db = get_db()
    return db.execute(
        """
        SELECT id, full_name, enrollment_no, department, program_name, semester
        FROM students
        ORDER BY full_name ASC
        """
    ).fetchall()


def fetch_cohort_breakdown(limit: int):
    db = get_db()
    return db.execute(
        """
        SELECT s.department,
               s.program_name,
               s.semester,
               COUNT(*) AS total_students,
               COUNT(a.id) AS present_today
        FROM students s
        LEFT JOIN attendance_logs a
               ON a.student_id = s.id
              AND a.attendance_date = ?
        GROUP BY s.department, s.program_name, s.semester
        ORDER BY total_students DESC, s.department ASC, s.program_name ASC, s.semester ASC
        LIMIT ?
        """,
        (date.today().isoformat(), limit),
    ).fetchall()


def fetch_schedule_for_today(limit: int):
    db = get_db()
    weekday = datetime.now().strftime("%A")
    return db.execute(
        """
        SELECT *
        FROM course_schedule
        WHERE weekday = ?
        ORDER BY start_time ASC
        LIMIT ?
        """,
        (weekday, limit),
    ).fetchall()


def fetch_schedule_for_student(department: str, program_name: str, semester: str, limit: int):
    db = get_db()
    weekday = datetime.now().strftime("%A")
    return db.execute(
        """
        SELECT *
        FROM course_schedule
        WHERE department = ?
          AND program_name = ?
          AND semester = ?
          AND weekday = ?
        ORDER BY start_time ASC
        LIMIT ?
        """,
        (department, program_name, semester, weekday, limit),
    ).fetchall()


def fetch_notices(limit: int, include_admin: bool = True):
    db = get_db()
    if include_admin:
        query = "SELECT * FROM notices ORDER BY created_at DESC LIMIT ?"
    else:
        query = "SELECT * FROM notices WHERE audience IN ('Students', 'Campus', 'All') ORDER BY created_at DESC LIMIT ?"
    return db.execute(query, (limit,)).fetchall()


def fetch_top_performers(limit: int):
    db = get_db()
    return db.execute(
        """
        SELECT s.full_name,
               s.enrollment_no,
               s.department,
               s.program_name,
               s.semester,
               ROUND(AVG((cr.score * 100.0) / cr.max_score), 1) AS average_percent
        FROM course_results cr
        JOIN students s ON s.id = cr.student_id
        GROUP BY s.id
        ORDER BY average_percent DESC, s.full_name ASC
        LIMIT ?
        """,
        (limit,),
    ).fetchall()


def fetch_course_performance(limit: int):
    db = get_db()
    return db.execute(
        """
        SELECT course_code,
               course_title,
               ROUND(AVG((score * 100.0) / max_score), 1) AS average_percent,
               COUNT(*) AS submissions
        FROM course_results
        GROUP BY course_code, course_title
        ORDER BY average_percent DESC, submissions DESC, course_title ASC
        LIMIT ?
        """,
        (limit,),
    ).fetchall()


def fetch_recent_results(limit: int, student_id: int | None = None):
    db = get_db()
    query = """
        SELECT cr.assessment_name,
               cr.course_code,
               cr.course_title,
               cr.score,
               cr.max_score,
               cr.grade,
               cr.updated_at,
               s.full_name,
               s.enrollment_no,
               s.department,
               s.program_name,
               s.semester
        FROM course_results cr
        JOIN students s ON s.id = cr.student_id
    """
    params = []
    if student_id is not None:
        query += " WHERE cr.student_id = ?"
        params.append(student_id)
    query += " ORDER BY cr.updated_at DESC LIMIT ?"
    params.append(limit)
    return db.execute(query, params).fetchall()


def get_attendance_metrics():
    metrics = get_dashboard_metrics()
    return {
        "total_students": metrics["total_students"],
        "present_today": metrics["present_today"],
        "pending_today": max(metrics["total_students"] - metrics["present_today"], 0),
        "attendance_rate": metrics["attendance_rate"],
    }


def fetch_pending_students(limit: int):
    db = get_db()
    return db.execute(
        """
        SELECT s.id,
               s.full_name,
               s.enrollment_no,
               s.department,
               s.program_name,
               s.semester
        FROM students s
        LEFT JOIN attendance_logs a
               ON a.student_id = s.id
              AND a.attendance_date = ?
        WHERE a.id IS NULL
        ORDER BY s.full_name ASC
        LIMIT ?
        """,
        (date.today().isoformat(), limit),
    ).fetchall()


def get_finance_summary():
    db = get_db()
    finance = db.execute(
        "SELECT COALESCE(SUM(amount_due), 0) AS total_due, COALESCE(SUM(amount_paid), 0) AS total_paid FROM fee_records"
    ).fetchone()
    overdue_count = db.execute(
        "SELECT COUNT(*) AS count FROM fee_records WHERE due_date < ? AND amount_paid < amount_due",
        (date.today().isoformat(),),
    ).fetchone()["count"]
    total_due = round(finance["total_due"], 2)
    total_paid = round(finance["total_paid"], 2)
    recovery_rate = round((total_paid * 100.0 / total_due), 1) if total_due else 0
    return {
        "total_due": total_due,
        "total_paid": total_paid,
        "outstanding": round(total_due - total_paid, 2),
        "recovery_rate": recovery_rate,
        "overdue_count": overdue_count,
    }


def fetch_fee_records(limit: int, student_id: int | None = None):
    db = get_db()
    query = """
        SELECT f.fee_head,
               f.semester_label,
               f.amount_due,
               f.amount_paid,
               f.status,
               f.due_date,
               f.updated_at,
               ROUND(f.amount_due - f.amount_paid, 2) AS balance_due,
               s.full_name,
               s.enrollment_no,
               s.department,
               s.program_name,
               s.semester
        FROM fee_records f
        JOIN students s ON s.id = f.student_id
    """
    params = []
    if student_id is not None:
        query += " WHERE f.student_id = ?"
        params.append(student_id)
    query += " ORDER BY f.updated_at DESC LIMIT ?"
    params.append(limit)
    return db.execute(query, params).fetchall()


def fetch_overdue_fees(limit: int):
    db = get_db()
    return db.execute(
        """
        SELECT f.fee_head,
               f.semester_label,
               f.amount_due,
               f.amount_paid,
               f.due_date,
               ROUND(f.amount_due - f.amount_paid, 2) AS balance_due,
               s.full_name,
               s.enrollment_no,
               s.program_name,
               s.semester
        FROM fee_records f
        JOIN students s ON s.id = f.student_id
        WHERE f.due_date < ?
          AND f.amount_paid < f.amount_due
        ORDER BY f.due_date ASC, balance_due DESC
        LIMIT ?
        """,
        (date.today().isoformat(), limit),
    ).fetchall()


def fetch_student_profile(student_id: int):
    db = get_db()
    return db.execute(
        """
        SELECT s.*,
               COALESCE(SUM(f.amount_due - f.amount_paid), 0) AS balance_due,
               COALESCE(
                   ROUND((SELECT AVG((cr.score * 100.0) / cr.max_score)
                          FROM course_results cr
                          WHERE cr.student_id = s.id), 1),
                   0
               ) AS academic_average
        FROM students s
        LEFT JOIN fee_records f ON f.student_id = s.id
        WHERE s.id = ?
        GROUP BY s.id
        """,
        (student_id,),
    ).fetchone()


def get_student_portal_metrics(student_id: int):
    db = get_db()
    today = date.today().isoformat()
    attendance_count = db.execute(
        "SELECT COUNT(*) AS count FROM attendance_logs WHERE student_id = ?",
        (student_id,),
    ).fetchone()["count"]
    today_mark = db.execute(
        "SELECT COUNT(*) AS count FROM attendance_logs WHERE student_id = ? AND attendance_date = ?",
        (student_id, today),
    ).fetchone()["count"]
    last_seen = db.execute(
        "SELECT logged_at FROM attendance_logs WHERE student_id = ? ORDER BY logged_at DESC LIMIT 1",
        (student_id,),
    ).fetchone()
    fee_row = db.execute(
        """
        SELECT COALESCE(SUM(amount_due), 0) AS total_due,
               COALESCE(SUM(amount_paid), 0) AS total_paid
        FROM fee_records
        WHERE student_id = ?
        """,
        (student_id,),
    ).fetchone()
    avg_score = db.execute(
        "SELECT ROUND(AVG((score * 100.0) / max_score), 1) AS avg_percent FROM course_results WHERE student_id = ?",
        (student_id,),
    ).fetchone()["avg_percent"]

    return {
        "attendance_marked": attendance_count,
        "present_today": bool(today_mark),
        "last_seen": last_seen["logged_at"] if last_seen else "Not marked yet",
        "academic_average": round(avg_score or 0, 1),
        "total_due": round(fee_row["total_due"], 2),
        "total_paid": round(fee_row["total_paid"], 2),
        "outstanding": round(fee_row["total_due"] - fee_row["total_paid"], 2),
    }


def get_current_student_or_none():
    user_id = session.get("user_id")
    if session.get("role") != "student" or not user_id:
        return None
    return fetch_student_profile(int(user_id))


@app.route("/")
def home():
    return redirect_for_role()


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        role = normalize_text(request.form.get("role", ""))
        username = normalize_text(request.form.get("username", ""))
        password = request.form.get("password", "")
        db = get_db()

        if role == "admin":
            admin = db.execute(
                "SELECT * FROM admin_users WHERE username = ?",
                (username,),
            ).fetchone()
            if admin and check_password_hash(admin["password_hash"], password):
                session.clear()
                session.update(
                    role="admin",
                    user_id=admin["id"],
                    display_name=admin["full_name"],
                    login_id=admin["username"],
                )
                return redirect(url_for("dashboard"))

        if role == "student":
            normalized_login = normalize_enrollment_no(username)
            student = db.execute(
                "SELECT * FROM students WHERE enrollment_no = ? COLLATE NOCASE",
                (normalized_login,),
            ).fetchone()
            if student and verify_student_password(student, password):
                session.clear()
                session.update(
                    role="student",
                    user_id=student["id"],
                    display_name=student["full_name"],
                    login_id=student["enrollment_no"],
                )
                return redirect(url_for("student_panel"))

        flash("Invalid credentials. Check your role, login ID, and password.", "error")

    if session.get("role"):
        return redirect_for_role()

    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    flash("You have been logged out.", "success")
    return redirect(url_for("login"))


@app.route("/dashboard")
@admin_required
def dashboard():
    return render_template(
        "dashboard.html",
        metrics=get_dashboard_metrics(),
        recent_attendance=fetch_recent_attendance(8),
        cohort_breakdown=fetch_cohort_breakdown(6),
        top_performers=fetch_top_performers(5),
        schedule_today=fetch_schedule_for_today(6),
        notices=fetch_notices(4, include_admin=True),
        finance_summary=get_finance_summary(),
    )


@app.route("/students")
@admin_required
def students():
    return render_template(
        "students.html",
        students=fetch_students(),
        cohort_breakdown=fetch_cohort_breakdown(6),
        metrics=get_dashboard_metrics(),
    )


@app.route("/employees")
@admin_required
def employees_redirect():
    return redirect(url_for("students"))


@app.route("/attendance")
@admin_required
def attendance():
    return render_template(
        "attendance.html",
        metrics=get_attendance_metrics(),
        recent_attendance=fetch_recent_attendance(12),
        cohort_breakdown=fetch_cohort_breakdown(8),
        pending_students=fetch_pending_students(8),
    )


@app.route("/academics")
@admin_required
def academics():
    return render_template(
        "academics.html",
        students=fetch_student_options(),
        metrics=get_dashboard_metrics(),
        top_performers=fetch_top_performers(6),
        course_performance=fetch_course_performance(6),
        recent_results=fetch_recent_results(12),
    )


@app.route("/finance")
@admin_required
def finance():
    return render_template(
        "finance.html",
        students=fetch_student_options(),
        summary=get_finance_summary(),
        fee_records=fetch_fee_records(12),
        overdue_fees=fetch_overdue_fees(8),
    )


@app.route("/student")
@app.route("/student/panel")
@student_required
def student_panel():
    student = get_current_student_or_none()
    if student is None:
        session.clear()
        flash("Student account was not found. Please log in again.", "error")
        return redirect(url_for("login"))

    return render_template(
        "student_panel.html",
        student=student,
        metrics=get_student_portal_metrics(student["id"]),
        attendance_history=fetch_recent_attendance(8, student["id"]),
        results=fetch_recent_results(8, student["id"]),
        fees=fetch_fee_records(8, student["id"]),
        notices=fetch_notices(4, include_admin=False),
        schedule=fetch_schedule_for_student(student["department"], student["program_name"], student["semester"], 6),
    )


@app.route("/api/students", methods=["POST"])
@admin_required
def create_student():
    enrollment_no = normalize_enrollment_no(request.form.get("enrollment_no", ""))
    full_name = normalize_text(request.form.get("full_name", ""))
    department = normalize_text(request.form.get("department", ""))
    program_name = normalize_text(request.form.get("program_name", ""))
    semester = normalize_text(request.form.get("semester", ""))
    batch_year = normalize_text(request.form.get("batch_year", ""))
    emergency_contact_name = normalize_text(request.form.get("emergency_contact_name", ""))
    emergency_contact_phone = normalize_text(request.form.get("emergency_contact_phone", ""))
    portal_password = request.form.get("portal_password", "")
    email = normalize_text(request.form.get("email", ""))
    dob = normalize_text(request.form.get("dob", ""))
    address = normalize_text(request.form.get("address", ""))
    captured_photo = request.form.get("captured_photo", "")
    photo = request.files.get("photo")

    required_values = {
        "enrollment number": enrollment_no,
        "full name": full_name,
        "department": department,
        "program": program_name,
        "semester": semester,
        "batch year": batch_year,
        "emergency contact name": emergency_contact_name,
        "emergency contact phone": emergency_contact_phone,
        "portal password": portal_password,
    }
    missing = [label for label, value in required_values.items() if not value]
    if missing:
        flash(f"Missing required fields: {', '.join(missing)}.", "error")
        return redirect(url_for("students"))

    image_path = ""
    try:
        image_path = save_face_image(photo, captured_photo, prefix=enrollment_no.replace(" ", "-").lower())
        face_encoding = serialize_face_encoding(extract_face_encoding(image_path=image_path))
    except ValueError as exc:
        if image_path:
            remove_saved_image(image_path)
        flash(str(exc), "error")
        return redirect(url_for("students"))

    db = get_db()
    try:
        db.execute(
            """
            INSERT INTO students (
                enrollment_no, full_name, department, program_name, semester, batch_year,
                emergency_contact_name, emergency_contact_phone, email, dob, address,
                status, image_path, face_encoding, portal_password_hash, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                enrollment_no,
                full_name,
                department,
                program_name,
                semester,
                batch_year,
                emergency_contact_name,
                emergency_contact_phone,
                email,
                dob,
                address,
                "Active",
                image_path,
                face_encoding,
                generate_password_hash(portal_password),
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            ),
        )
        db.commit()
    except sqlite3.IntegrityError:
        remove_saved_image(image_path)
        flash("Enrollment number already exists. Use a unique ID.", "error")
        return redirect(url_for("students"))

    flash("College student profile created and facial recognition is ready.", "success")
    return redirect(url_for("students"))


@app.route("/api/students/reset-password", methods=["POST"])
@admin_required
def reset_student_password():
    student_id = request.form.get("student_id", type=int)
    new_password = request.form.get("new_password", "")
    if not student_id:
        flash("Student record was not found.", "error")
        return redirect(url_for("students"))
    if len(new_password.strip()) < 6:
        flash("Portal password must be at least 6 characters.", "error")
        return redirect(url_for("students"))

    db = get_db()
    student = db.execute(
        "SELECT full_name, enrollment_no FROM students WHERE id = ?",
        (student_id,),
    ).fetchone()
    if student is None:
        flash("Student record was not found.", "error")
        return redirect(url_for("students"))

    db.execute(
        "UPDATE students SET portal_password_hash = ? WHERE id = ?",
        (generate_password_hash(new_password), student_id),
    )
    db.commit()
    flash(
        f"Portal password reset for {student['full_name']} ({student['enrollment_no']}).",
        "success",
    )
    return redirect(url_for("students"))


@app.route("/api/academics/results", methods=["POST"])
@admin_required
def add_result():
    student_id = request.form.get("student_id", type=int)
    assessment_name = normalize_text(request.form.get("assessment_name", ""))
    course_code = normalize_course_code(request.form.get("course_code", ""))
    course_title = normalize_text(request.form.get("course_title", ""))
    score = request.form.get("score", type=float)
    max_score = request.form.get("max_score", type=float)

    if not all([student_id, assessment_name, course_code, course_title]) or score is None or max_score is None:
        flash("Fill in all course result fields before saving.", "error")
        return redirect(url_for("academics"))

    if max_score <= 0 or score < 0 or score > max_score:
        flash("Score must be between 0 and max score.", "error")
        return redirect(url_for("academics"))

    db = get_db()
    upsert_course_results(
        db,
        [
            {
                "student_id": student_id,
                "assessment_name": assessment_name,
                "course_code": course_code,
                "course_title": course_title,
                "score": score,
                "max_score": max_score,
            }
        ],
    )
    db.commit()

    flash("Course result saved or updated in the academic ledger.", "success")
    return redirect(url_for("academics"))


@app.route("/api/academics/results/upload", methods=["POST"])
@admin_required
def bulk_upload_results():
    results_file = request.files.get("results_file")
    if results_file is None or not results_file.filename:
        flash("Choose an Excel or CSV marks file before uploading.", "error")
        return redirect(url_for("academics"))

    try:
        parsed_rows, skipped_blank_rows = parse_result_upload(results_file)
    except ValueError as exc:
        flash(str(exc), "error")
        return redirect(url_for("academics"))

    if not parsed_rows:
        flash("No marks rows were found in the uploaded file.", "error")
        return redirect(url_for("academics"))

    db = get_db()
    student_map = {
        normalize_enrollment_no(row["enrollment_no"]): row["id"]
        for row in db.execute("SELECT id, enrollment_no FROM students").fetchall()
    }

    prepared_rows = {}
    missing_students = []
    for row in parsed_rows:
        student_id = student_map.get(row["enrollment_no"])
        if student_id is None:
            missing_students.append(row["enrollment_no"])
            continue
        key = (student_id, row["assessment_name"], row["course_code"])
        prepared_rows[key] = {
            "student_id": student_id,
            **row,
        }

    if not prepared_rows:
        flash("No matching students were found for the uploaded marks file.", "error")
        return redirect(url_for("academics"))

    upserted_count = upsert_course_results(db, list(prepared_rows.values()))
    db.commit()

    message_parts = [f"Upserted {upserted_count} marks rows."]
    if skipped_blank_rows:
        message_parts.append(f"Skipped {skipped_blank_rows} blank rows.")
    if missing_students:
        unique_missing = sorted(set(missing_students))
        preview = ", ".join(unique_missing[:5])
        suffix = "..." if len(unique_missing) > 5 else ""
        message_parts.append(f"Missing students: {preview}{suffix}")

    flash(" ".join(message_parts), "success")
    return redirect(url_for("academics"))


@app.route("/api/finance/fees", methods=["POST"])
@admin_required
def add_fee_record():
    student_id = request.form.get("student_id", type=int)
    fee_head = normalize_text(request.form.get("fee_head", ""))
    semester_label = normalize_text(request.form.get("semester_label", ""))
    due_date = normalize_text(request.form.get("due_date", ""))
    amount_due = request.form.get("amount_due", type=float)
    amount_paid = request.form.get("amount_paid", type=float)

    if not all([student_id, fee_head, semester_label, due_date]) or amount_due is None or amount_paid is None:
        flash("Fill in all fee fields before saving.", "error")
        return redirect(url_for("finance"))

    if amount_due < 0 or amount_paid < 0:
        flash("Fee amounts cannot be negative.", "error")
        return redirect(url_for("finance"))

    db = get_db()
    db.execute(
        """
        INSERT INTO fee_records (
            student_id, fee_head, semester_label, amount_due, amount_paid,
            status, due_date, updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            student_id,
            fee_head,
            semester_label,
            amount_due,
            amount_paid,
            fee_status(amount_due, amount_paid, due_date),
            due_date,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        ),
    )
    db.commit()

    flash("Fee ledger updated for the selected student.", "success")
    return redirect(url_for("finance"))


@app.route("/api/attendance/recognize", methods=["POST"])
@admin_required
def recognize_attendance():
    payload = request.get_json(silent=True) or {}
    image_data = payload.get("image", "")
    if not image_data:
        return jsonify({"ok": False, "message": "Attendance image is missing."}), 400

    try:
        target_encoding = extract_face_encoding(image_data_url=image_data)
    except ValueError as exc:
        return jsonify({"ok": False, "message": str(exc)}), 400

    db = get_db()
    students = db.execute(
        """
        SELECT id, full_name, enrollment_no, department, program_name, semester, face_encoding
        FROM students
        WHERE face_encoding IS NOT NULL AND face_encoding != ''
        """
    ).fetchall()

    if not students:
        return jsonify({"ok": False, "message": "No enrolled student profiles are ready for recognition yet."}), 400

    known_encodings = [parse_face_encoding(row["face_encoding"]) for row in students]
    if HAS_FACE_RECOGNITION and face_recognition is not None:
        distances = face_recognition.face_distance(known_encodings, target_encoding)
    else:
        distances = np.linalg.norm(np.array(known_encodings) - target_encoding, axis=1)

    best_index = int(np.argmin(distances))
    best_distance = float(distances[best_index])


    if best_distance > MATCH_THRESHOLD:
        return jsonify(
            {
                "ok": False,
                "message": "Face not recognized. Capture a clearer image or enroll the student first.",
            }
        ), 404

    matched = students[best_index]
    confidence = round(max(0.0, min(100.0, (1.0 - best_distance) * 100.0)), 1)
    today = date.today().isoformat()
    existing = db.execute(
        "SELECT id FROM attendance_logs WHERE student_id = ? AND attendance_date = ?",
        (matched["id"], today),
    ).fetchone()

    already_marked = existing is not None
    if not already_marked:
        db.execute(
            """
            INSERT INTO attendance_logs (student_id, attendance_date, status, confidence, logged_at)
            VALUES (?, ?, ?, ?, ?)
            """,
            (
                matched["id"],
                today,
                "Present",
                confidence,
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            ),
        )
        db.commit()

    cohort = f"{matched['department']} | {matched['program_name']} | {matched['semester']}"
    return jsonify(
        {
            "ok": True,
            "already_marked": already_marked,
            "confidence": confidence,
            "student": {
                "name": matched["full_name"],
                "code": matched["enrollment_no"],
                "cohort": cohort,
            },
        }
    )


init_db()


if __name__ == "__main__":
    app.run(debug=True)
